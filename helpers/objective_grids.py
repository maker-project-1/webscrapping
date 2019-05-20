from copy import deepcopy

import pandas as pd
from openpyxl import load_workbook

obj_grid_2017_paths = [
    '/home/renaud/Dropbox/Etaonis/Google Drive/Missions/MH/Scraping/Objectives Grids 2017/Objective Grid - USA - Central.xlsx',
    '/home/renaud/Dropbox/Etaonis/Google Drive/Missions/MH/Scraping/Objectives Grids 2017/Objective Grid - USA - Nationwide.xlsx',
    '/home/renaud/Dropbox/Etaonis/Google Drive/Missions/MH/Scraping/Objectives Grids 2017/Objective Grid - USA - Northeast.xlsx',
    '/home/renaud/Dropbox/Etaonis/Google Drive/Missions/MH/Scraping/Objectives Grids 2017/Objective Grid - USA - West.xlsx',
    '/home/renaud/Dropbox/Etaonis/Google Drive/Missions/MH/Scraping/Objectives Grids 2017/Objective Grid - USA - Southeast.xlsx',
]

list_of_action_types = ['Product Image', 'Listed Range', 'Product Name']
list_of_status = ['target', 'plan']
cols_order = ['shop_id', 'brnd', 'pdct_name', 'action_type', 'target', 'plan']

l = []
for obj_grid_2017_path in obj_grid_2017_paths:
    print(obj_grid_2017_path)
    wb = load_workbook(obj_grid_2017_path, data_only=True)
    ws1 = wb.active

    retailers = [ws1.cell(row=16, column=9 + x*2).internal_value for x in range(50)]
    retailers = list(filter(None.__ne__, retailers))
    retailers = [x for pair in zip(retailers, retailers) for x in pair]

    for row_cells in ws1.iter_rows(row_offset=18):
        if row_cells[1].value not in list_of_action_types:
            continue
        brnd = row_cells[0].value
        action_type = row_cells[1].value
        pdct_name = row_cells[2].value
        pdct_name = " ".join(pdct_name.split()).replace('–', '-').replace("Demi-sec", "Demi sec")
        tmp_retailers = deepcopy(retailers)
        tmp_shop_id = ''
        for c, cell in enumerate(row_cells[8:]):
            if cell.value is not None:
                # print(c, cell.value)
                shop_id = tmp_retailers.pop(0)
                if tmp_shop_id != shop_id:
                    tmp={'shop_id': shop_id, 'brnd': brnd, 'action_type': action_type,
                              'pdct_name': pdct_name, list_of_status[c % 2]: cell.value}
                    tmp_shop_id = shop_id
                else:
                    val = cell.value
                    tmp.update({list_of_status[c % 2]: cell.value})
                    l.append(tmp)
pd.DataFrame(l)[cols_order].to_csv('/tmp/Agregated Objective Grids 2017.csv', sep=';', index=None)